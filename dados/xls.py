import openpyxl


wb = openpyxl.load_workbook("M5_02_Excel_data.xlsx")

print("Nomes de folhas: ")
print(wb.sheetnames)

print("\nNomes da folhas (outra folha): ")
for sheet in wb:
    print(sheet.title)

folha_um = wb.sheetnames[0]
print("\nPrimeira folha: ")
print(folha_um)