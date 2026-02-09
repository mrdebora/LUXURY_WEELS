import openpyxl
from openpyxl import Workbook


wb = Workbook()


ws = wb.active
ws.title = "Dados"


ws['A1'] = "Nome"
ws['B1'] = "Idade"
ws['C1'] = "Cidade"

ws['A2'] = "João"
ws['B2'] = 25
ws['C2'] = "São Paulo"

ws['A3'] = "Maria"
ws['B3'] = 30
ws['C3'] = "Rio de Janeiro"

ws['A4'] = "Pedro"
ws['B4'] = 28
ws['C4'] = "Belo Horizonte"


ws2 = wb.create_sheet("Vendas")
ws2['A1'] = "Produto"
ws2['B1'] = "Quantidade"
ws2['C1'] = "Preço"

ws2['A2'] = "Notebook"
ws2['B2'] = 5
ws2['C2'] = 2500.00

ws2['A3'] = "Mouse"
ws2['B3'] = 20
ws2['C3'] = 50.00


wb.save("M5_02_Excel_data.xlsx")
print("Arquivo Excel criado com sucesso!")
